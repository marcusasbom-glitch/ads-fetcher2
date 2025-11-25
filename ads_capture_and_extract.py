# ads_capture_and_extract.py – iframes + smart bildval + OCR från annonsbilder

import asyncio
import json
import os
from io import BytesIO
from pathlib import Path

import pandas as pd
import requests
import pytesseract
from PIL import Image
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright

# ============================================================
# Dynamiska paths (Render skapar en unik run-mapp per jobb)
# ============================================================

OUTPUT_DIR = Path("network_dump")
CANDIDATES_PATH = OUTPUT_DIR / "ads_candidates.json"
IMAGES_DIR = Path("images")
OUTPUT_EXCEL = "ads_extracted.xlsx"

# ta upp till 500 annonser
MAX_ADS = int(os.getenv("MAX_ADS", "1000"))
# hur många annonser som vi får köra OCR på (resten hoppar över OCR)
MAX_OCR_ADS = int(os.getenv("MAX_OCR_ADS", "150"))
DOWNLOAD_IMAGES = os.getenv("DOWNLOAD_IMAGES", "1") not in ("0", "false", "False")



def set_paths(base_dir):
    """Repoint global paths into the run dir."""
    global OUTPUT_DIR, CANDIDATES_PATH, IMAGES_DIR, OUTPUT_EXCEL

    base = Path(base_dir)
    OUTPUT_DIR = base / "network_dump"
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    CANDIDATES_PATH = base / "ads_candidates.json"

    IMAGES_DIR = base / "images"
    IMAGES_DIR.mkdir(parents=True, exist_ok=True)

    OUTPUT_EXCEL = str(base / "ads_extracted.xlsx")


# ============================================================
# Hygienfunktioner
# ============================================================

def sanitize_filename(name):
    import re
    return re.sub(r"[^a-zA-Z0-9._-]", "_", name)


def get_available_filename(base):
    p = Path(base)
    if not p.exists():
        return str(p)
    stem = p.stem
    ext = p.suffix
    for i in range(1, 9999):
        alt = p.with_name(f"{stem}_{i}{ext}")
        if not alt.exists():
            return str(alt)
    return str(p)


def ocr_image(path: str) -> str:
    """Läs text ur en bild med Tesseract (svenska + engelska)."""
    try:
        img = Image.open(path)
    except Exception as e:
        print(f"⚠️ OCR: kunde inte öppna bild {path}: {e}")
        return ""
    try:
        return pytesseract.image_to_string(img, lang="swe+eng")
    except Exception as e:
        print(f"⚠️ OCR swe+eng misslyckades ({path}): {e}")
        try:
            return pytesseract.image_to_string(img)
        except Exception as e2:
            print(f"⚠️ OCR utan språk misslyckades ({path}): {e2}")
            return ""


# ============================================================
# PLAYWRIGHT – DOM scraping i ALLA FRAMES
# ============================================================

# JavaScript-kod som körs inne i varje frame
# Vi kräver minst en "stor" bild, och sparar alla textrader.
IFRAME_JS_SCRAPER = """
() => {
    const cards = [];
    const root = document.body;
    if (!root) return cards;

    const MIN_AREA = 10000; // 100x100
    const elements = root.querySelectorAll("article, div, section");

    for (const el of elements) {
        const txt = (el.innerText || "").trim();
        const imgNodes = Array.from(el.querySelectorAll("img"));
        if (!imgNodes.length) continue;

        const imgInfos = imgNodes.map(i => {
            const rect = i.getBoundingClientRect();
            const w = i.naturalWidth || i.width || rect.width || 0;
            const h = i.naturalHeight || i.height || rect.height || 0;
            const area = w * h;
            const ratio = h > 0 ? (w / h) : 0;
            return {
                src: i.src,
                w: w,
                h: h,
                area: area,
                top: rect.top || 0,
                left: rect.left || 0,
                ratio: ratio
            };
        });

        const hasBig = imgInfos.some(info => info.area >= MIN_AREA);
        if (!hasBig) continue;

        const lines = (txt || "").split("\\n").map(s => s.trim()).filter(Boolean);

        let headNode =
            el.querySelector("h1, h2, h3, h4") ||
            el.querySelector('a[role="heading"]') ||
            el.querySelector("a");

        let headline = headNode ? (headNode.innerText || "").trim() : "";
        if (!headline && lines.length > 1) {
            headline = lines[1];
        }

        const advertiser = lines.length ? lines[0] : "";

        cards.push({
            advertiser: advertiser,
            headline: headline,
            text: txt,
            lines: lines,
            image_infos: imgInfos
        });
    }

    return cards;
}
"""


async def capture_network(ar_input, run_dir):
    """Scrape ALLA frames (iframes) på Google Ads Transparency-sidan."""

    set_paths(run_dir)

    if ar_input.startswith("http"):
        url = ar_input
    else:
        url = (
            "https://adstransparency.google.com/advertiser/"
            + ar_input +
            "?origin=ata&region=SE&preset-date=Last+7+days&platform=SEARCH"
        )

    print("🔗 Laddar:", url)

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        ctx = await browser.new_context(
            user_agent="Mozilla/5.0",
            locale="sv-SE"
        )
        page = await ctx.new_page()

        await page.goto(url, wait_until="domcontentloaded", timeout=45000)

        # Scroll i huvudsidan (om något laddas där)
        for _ in range(10):
            await page.evaluate("window.scrollBy(0, window.innerHeight)")
            await asyncio.sleep(0.6)

        all_ads = []

        # Huvud-frame
        try:
            main_ads = await page.evaluate(IFRAME_JS_SCRAPER)
            print(f"🧩 Huvud-frame: {len(main_ads)} annonskort")
            all_ads.extend(main_ads)
        except Exception as e:
            print("Fel vid DOM-scrape i huvud-frame:", e)

        # Alla övriga frames – scrolla och scrapa
        for frame in page.frames:
            if frame == page.main_frame:
                continue
            try:
                # scrolla inne i iframen för att trigga lazy load
                for _ in range(10):
                    await frame.evaluate("window.scrollBy(0, window.innerHeight)")
                    await asyncio.sleep(0.4)

                frame_ads = await frame.evaluate(IFRAME_JS_SCRAPER)
                print(f"🪟 Frame {frame.url} → {len(frame_ads)} annonskort")
                all_ads.extend(frame_ads)
            except Exception as e:
                print("⚠️ Frame ej läsbar (cross-origin):", frame.url, e)

        print("✅ TOTALT hittade annonskort:", len(all_ads))

        OUTPUT_DIR.mkdir(exist_ok=True, parents=True)
        CANDIDATES_PATH.write_text(
            json.dumps(
                [{"source_file": "frames_dom", "parsed": all_ads}],
                indent=2, ensure_ascii=False
            ),
            encoding="utf-8"
        )

        await browser.close()
        return True


# ============================================================
# EXCEL + bildnedladdning + OCR
# ============================================================

def process_candidates_and_save(run_dir):
    set_paths(run_dir)

    if not CANDIDATES_PATH.exists():
        print("❌ Saknar ads_candidates.json")
        return False

    data = json.loads(CANDIDATES_PATH.read_text(encoding="utf-8"))
    if not data:
        print("❌ Tomma kandidater")
        return False

    ads = data[0].get("parsed") or []
    print("⏳ Bearbetar annonser:", len(ads))

    rows = []
    session = requests.Session()
    session.headers.update({"User-Agent": "Mozilla/5.0"})

    for idx, ad in enumerate(ads, start=1):
        if idx > MAX_ADS:
            break

        infos = ad.get("image_infos") or []

        # välj "bästa" bild:
        big = [i for i in infos if (i.get("area") or 0) >= 10000]
        if not big:
            big = infos

        best = None
        best_score = -1.0
        for info in big:
            try:
                area = float(info.get("area") or 0)
                ratio = float(info.get("ratio") or 0)
            except Exception:
                area = 0.0
                ratio = 0.0
            score = area
            if ratio > 1.2:
                score *= 1.5  # bonus för landskap
            if score > best_score:
                best_score = score
                best = info

        img_url = best["src"] if best and best.get("src") else ""
        img_file = ""

        if img_url and DOWNLOAD_IMAGES:
            try:
                url = img_url
                if url.startswith("//"):
                    url = "https:" + url

                r = session.get(url, timeout=10)
                if r.status_code == 200:
                    ct = (r.headers.get("content-type") or "").lower()
                    ext = "png"
                    if "jpg" in ct or "jpeg" in ct:
                        ext = "jpg"
                    elif "webp" in ct:
                        ext = "webp"

                    fname = sanitize_filename(f"ad_{idx}.{ext}")
                    path = IMAGES_DIR / fname
                    path.write_bytes(r.content)
                    img_file = str(path)
            except Exception as e:
                print("⚠️ kunde inte ladda ner bild:", img_url, e)
                img_file = ""

        # bygg beskrivning från text-rader efter rubrik
        lines = ad.get("lines") or []
        advertiser = ad.get("advertiser", "")
        headline = ad.get("headline", "")
        description = ""
        if len(lines) >= 3:
            description = " ".join(lines[2:])
        elif len(lines) == 2:
            description = lines[1]

        # OCR-fallback om vi saknar rubrik + beskrivning men har bild
                # kör bara OCR för de första MAX_OCR_ADS annonserna
        do_ocr = (idx <= MAX_OCR_ADS)

        # OCR-fallback om vi saknar rubrik + beskrivning men har bild
        if do_ocr and not headline.strip() and not description.strip() and img_file:
            ocr_txt = ocr_image(img_file)
            ocr_lines = [l.strip() for l in ocr_txt.splitlines() if l.strip()]
            if ocr_lines:
                if not headline.strip():
                    headline = ocr_lines[0][:120]
                if len(ocr_lines) > 1 and not description.strip():
                    description = " ".join(ocr_lines[1:])[:500]


        rows.append({
            "Index": idx,
            "Annonsör": advertiser,
            "Rubrik": headline,
            "Beskrivning": description,
            "Text": ad.get("text", ""),
            "Bild-URL": img_url,
            "Bildfil": img_file,
        })

    if not rows:
        df = pd.DataFrame([{"Info": "Inga annonser hittades"}])
        excel = get_available_filename(OUTPUT_EXCEL)
        df.to_excel(excel, index=False)
        print("📄 Excel utan annonser:", excel)
        return True

    df = pd.DataFrame(rows)
    excel = get_available_filename(OUTPUT_EXCEL)
    df.to_excel(excel, index=False)
    print("📊 Grund-Excel sparad:", excel)

    # Bädda in bilder
    wb = load_workbook(excel)
    ws = wb.active

    for r, row in enumerate(rows, start=2):
        f = row["Bildfil"]
        if not f or not Path(f).exists():
            continue

        try:
            img = Image.open(f)
            w, h = img.size
            scale = min(200 / w, 200 / h, 1.0)
            if scale < 1.0:
                img = img.resize((int(w * scale), int(h * scale)))

            bio = BytesIO()
            img.save(bio, format="PNG")
            bio.seek(0)

            xlimg = XLImage(bio)
            xlimg.width = 140
            xlimg.height = 140
            ws.add_image(xlimg, f"G{r}")  # kolumn G = Bildfil
            ws.row_dimensions[r].height = 110
        except Exception as e:
            print("Fel vid inbäddning av bild på rad", r, ":", e)

    # snygga kolumner
    for i, col in enumerate(df.columns, start=1):
        col_letter = get_column_letter(i)
        maxlen = max((len(str(x)) for x in df[col]), default=len(col))
        ws.column_dimensions[col_letter].width = min(maxlen + 6, 80)

    wb.save(excel)
    print("✅ Excel med inbäddade bilder:", excel)
    return True



