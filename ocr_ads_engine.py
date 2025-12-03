import pytesseract
from PIL import Image
from openpyxl import load_workbook
from pathlib import Path


def add_ocr_to_excel(input_xlsx: Path, output_xlsx: Path):
    """
    Läser en Excel-fil från din scraper där annonsbilder redan ligger som PNG i tempmappen.
    Kör OCR på varje bild, lägger till kolumner:
        OCR_Rubrik
        OCR_Beskrivning
    och sparar en ny Excel-fil.
    """

    wb = load_workbook(input_xlsx)
    ws = wb.active

    # Hitta kolumner för Bildfil
    imagefile_col = None
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header and header.lower().strip() == "bildfil":
            imagefile_col = col
            break

    if not imagefile_col:
        raise RuntimeError("Kunde inte hitta kolumnen 'Bildfil' i Excel-filen.")

    # Skapa två nya kolumner
    ocr_title_col = ws.max_column + 1
    ws.cell(row=1, column=ocr_title_col, value="OCR_Rubrik")

    ocr_desc_col = ws.max_column + 1
    ws.cell(row=1, column=ocr_desc_col, value="OCR_Beskrivning")

    for row in range(2, ws.max_row + 1):
        img_path = ws.cell(row=row, column=imagefile_col).value
        if not img_path or not Path(img_path).exists():
            continue

        try:
            img = Image.open(img_path)
            text = pytesseract.image_to_string(img)

            # enkel delning i första raden + resten
            lines = [l.strip() for l in text.splitlines() if l.strip()]
            if lines:
                title = lines[0]
                desc = " ".join(lines[1:])
            else:
                title = ""
                desc = ""

            ws.cell(row=row, column=ocr_title_col).value = title
            ws.cell(row=row, column=ocr_desc_col).value = desc

        except Exception as e:
            ws.cell(row=row, column=ocr_title_col).value = f"OCR error: {e}"

    wb.save(output_xlsx)
